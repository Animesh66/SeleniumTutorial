import openpyxl as xl

path = "/Users/animeshmukherjee/Desktop/Animesh/transactions.xlsx"
workbook = xl.load_workbook(path)  # this command will load the workbook from the location
sheet = workbook["Sheet1"]  # this will select the sheet in the workbook
rows = sheet.max_row  # this will give the total number of row
columns = sheet.max_column  # this will give the total number of columns
for row in range(1, rows + 1):  # traverse from 1st row to last row
    for column in range(1, columns + 1):  # traverse from 1st column to last column
        cell = sheet.cell(row, column)  # this will return the current cell
        print(cell.value)  # it will print the value of the cell from the excel

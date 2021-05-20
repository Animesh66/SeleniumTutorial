import openpyxl as xl


def get_row_count(file,sheet_name):  # this function will return the row count of the cell
    workbook = xl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(file, sheet_name):  # this function will return the column count of the cell
    workbook = xl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(file, sheet_name, row_count, column_count):  #  this function will read the value of the cell from excel file
    workbook = xl.load_workbook(file)
    sheet = workbook[sheet_name]
    cell = sheet.cell(row_count,column_count)
    return cell.value


def write_data(file, sheet_name, row_count, column_count, cell_value):  # this function will write data in excel
    workbook = xl.load_workbook(file)
    sheet = workbook[sheet_name]
    cell = sheet.cell(row_count, column_count)
    cell.value = cell_value
    workbook.save(file)
